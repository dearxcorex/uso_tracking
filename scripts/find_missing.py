"""
Find data in update.xlsx and school.xlsx that is MISSING from preview_import.xlsx.
Uses (latitude, longitude) rounded to 5 decimal places as primary match key,
with install_location text as fallback.
"""

import openpyxl
from collections import defaultdict


def round_coord(val, decimals=5):
    """Round a coordinate value, return None if not numeric."""
    if val is None:
        return None
    try:
        return round(float(val), decimals)
    except (ValueError, TypeError):
        return None


def normalize_text(text):
    """Normalize text for comparison: strip whitespace, lowercase."""
    if text is None:
        return ""
    return str(text).strip().lower()


def load_preview_import(path="data/preview_import.xlsx"):
    """Load the main reference file. Returns set of (lat, lng) keys and set of install_location keys."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    coord_keys = set()
    location_keys = set()
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        lat = round_coord(row[7])
        lng = round_coord(row[8])
        install_loc = normalize_text(row[5])

        if lat is not None and lng is not None:
            coord_keys.add((lat, lng))
        if install_loc:
            location_keys.add(install_loc)

        rows.append({
            "service_name": row[0],
            "village": row[1],
            "subdistrict": row[2],
            "district": row[3],
            "province": row[4],
            "install_location": row[5],
            "provider": row[6],
            "latitude": row[7],
            "longitude": row[8],
            "zone": row[9],
            "contract_number": row[10],
        })

    wb.close()
    print(f"[preview_import.xlsx] Loaded {len(rows)} rows, {len(coord_keys)} unique coord keys, {len(location_keys)} location keys")
    return coord_keys, location_keys, rows


def load_update_sheets(path="data/update.xlsx"):
    """Load data sheets from update.xlsx, deduplicating across sheets by coordinates."""
    wb = openpyxl.load_workbook(path, read_only=True)

    # Data sheets to read (skip summary sheets)
    data_sheets = ["ชัยภูมิ", "กันไฟฟ้าดูด", "mobile-wiFi Calling", "USO Net-USO Wrap"]
    all_rows = []
    seen_coords = set()
    total_raw = 0
    dupes = 0

    for sheet_name in data_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: sheet '{sheet_name}' not found in update.xlsx")
            continue

        ws = wb[sheet_name]
        count = 0
        sheet_dupes = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            total_raw += 1

            lat = round_coord(row[10])
            lng = round_coord(row[11])
            coord_key = (lat, lng) if lat is not None and lng is not None else None

            # Deduplicate across sheets
            if coord_key and coord_key in seen_coords:
                sheet_dupes += 1
                dupes += 1
                continue
            if coord_key:
                seen_coords.add(coord_key)

            count += 1
            all_rows.append({
                "source_sheet": sheet_name,
                "ลำดับ": row[0],
                "รหัสหมู่บ้าน": row[1],
                "ประเภทบริการ": row[2],
                "ชื่อบริการ": row[3],
                "หมู่บ้าน": row[4],
                "ตำบล": row[5],
                "อำเภอ": row[6],
                "จังหวัด": row[7],
                "สถานที่ติดตั้ง": row[8],
                "ผู้ให้บริการ": row[9],
                "latitude": row[10],
                "longitude": row[11],
                "โครงการ": row[12],
            })
        print(f"  [update.xlsx / {sheet_name}] Loaded {count} rows (skipped {sheet_dupes} cross-sheet duplicates)")

    wb.close()
    print(f"[update.xlsx] Total: {len(all_rows)} unique rows ({total_raw} raw, {dupes} cross-sheet duplicates removed)")
    return all_rows


def load_school_sheets(path="data/school.xlsx"):
    """Load data sheets from school.xlsx, filtering for ชัยภูมิ province only."""
    wb = openpyxl.load_workbook(path, read_only=True)

    data_sheets = ["WIFI C", "WIFI C+", "WIFI C โรงเรียน", "WIFI C+ โรงเรียน"]
    all_rows = []

    for sheet_name in data_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: sheet '{sheet_name}' not found in school.xlsx")
            continue

        ws = wb[sheet_name]
        count = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            # Filter for ชัยภูมิ only
            province = str(row[7]).strip() if row[7] else ""
            if "ชัยภูมิ" not in province:
                skipped += 1
                continue
            count += 1
            all_rows.append({
                "source_sheet": sheet_name,
                "ลำดับ": row[0],
                "รหัสหมู่บ้าน": row[1],
                "ประเภทบริการ": row[2],
                "ชื่อบริการ": row[3],
                "หมู่บ้าน": row[4],
                "ตำบล": row[5],
                "อำเภอ": row[6],
                "จังหวัด": row[7],
                "สถานที่ติดตั้ง": row[8],
                "ผู้ให้บริการ": row[9],
                "latitude": row[10],
                "longitude": row[11],
                "โครงการ": row[12],
            })
        print(f"  [school.xlsx / {sheet_name}] Loaded {count} ชัยภูมิ rows (skipped {skipped} other provinces)")

    wb.close()
    print(f"[school.xlsx] Total: {len(all_rows)} ชัยภูมิ rows from {len(data_sheets)} sheets")
    return all_rows


def find_missing(source_rows, coord_keys, location_keys, source_name):
    """Find rows from source that are NOT in preview_import (by coord or install_location)."""
    missing = []
    matched_by_coord = 0
    matched_by_location = 0

    for row in source_rows:
        lat = round_coord(row["latitude"])
        lng = round_coord(row["longitude"])
        install_loc = normalize_text(row.get("สถานที่ติดตั้ง"))

        # Primary match: coordinates
        if lat is not None and lng is not None and (lat, lng) in coord_keys:
            matched_by_coord += 1
            continue

        # Fallback: install_location text
        if install_loc and install_loc in location_keys:
            matched_by_location += 1
            continue

        missing.append(row)

    print(f"\n[{source_name}] Match results:")
    print(f"  Total rows: {len(source_rows)}")
    print(f"  Matched by coordinates: {matched_by_coord}")
    print(f"  Matched by install_location: {matched_by_location}")
    print(f"  MISSING from preview_import: {len(missing)}")

    return missing


def print_summary(missing_rows, source_name):
    """Print summary breakdown of missing rows."""
    if not missing_rows:
        print(f"\n{'='*60}")
        print(f"No missing rows from {source_name}!")
        return

    print(f"\n{'='*60}")
    print(f"MISSING from preview_import.xlsx ({source_name}): {len(missing_rows)} rows")
    print(f"{'='*60}")

    # By service type
    by_service = defaultdict(int)
    for row in missing_rows:
        service = row.get("ชื่อบริการ") or "ไม่ระบุ"
        by_service[str(service).strip()] += 1

    print(f"\nBy service type (ชื่อบริการ):")
    for svc, count in sorted(by_service.items(), key=lambda x: -x[1]):
        print(f"  {svc}: {count}")

    # By district
    by_district = defaultdict(int)
    for row in missing_rows:
        district = row.get("อำเภอ") or "ไม่ระบุ"
        by_district[str(district).strip()] += 1

    print(f"\nBy district (อำเภอ):")
    for dist, count in sorted(by_district.items(), key=lambda x: -x[1]):
        print(f"  {dist}: {count}")

    # By source sheet
    by_sheet = defaultdict(int)
    for row in missing_rows:
        by_sheet[row["source_sheet"]] += 1

    print(f"\nBy source sheet:")
    for sheet, count in sorted(by_sheet.items(), key=lambda x: -x[1]):
        print(f"  {sheet}: {count}")

    # By zone
    by_zone = defaultdict(int)
    for row in missing_rows:
        zone = row.get("โครงการ") or "ไม่ระบุ"
        by_zone[str(zone).strip()] += 1

    print(f"\nBy zone (โครงการ):")
    for zone, count in sorted(by_zone.items(), key=lambda x: -x[1]):
        print(f"  {zone}: {count}")

    # Sample rows
    print(f"\nSample missing rows (first 10):")
    print(f"{'─'*120}")
    for i, row in enumerate(missing_rows[:10]):
        print(
            f"  {i+1}. [{row['source_sheet']}] "
            f"{row.get('ชื่อบริการ', '')!s:20s} | "
            f"{row.get('หมู่บ้าน', '')!s:30s} | "
            f"{row.get('อำเภอ', '')!s:20s} | "
            f"lat={row.get('latitude')}, lng={row.get('longitude')}"
        )


def check_duplicates_between_sources(update_missing, school_missing):
    """Check overlap between update and school missing rows."""
    update_coords = set()
    for row in update_missing:
        lat = round_coord(row["latitude"])
        lng = round_coord(row["longitude"])
        if lat is not None and lng is not None:
            update_coords.add((lat, lng))

    overlap = 0
    for row in school_missing:
        lat = round_coord(row["latitude"])
        lng = round_coord(row["longitude"])
        if lat is not None and lng is not None and (lat, lng) in update_coords:
            overlap += 1

    print(f"\n{'='*60}")
    print(f"OVERLAP between update.xlsx and school.xlsx missing rows: {overlap}")
    print(f"Unique missing from update.xlsx only: {len(update_missing) - overlap} (approx)")
    print(f"Unique missing from school.xlsx only: {len(school_missing) - overlap} (approx)")
    print(f"Total unique missing (approx): {len(update_missing) + len(school_missing) - overlap}")


def main():
    print("=" * 60)
    print("Finding data MISSING from preview_import.xlsx")
    print("=" * 60)

    # Step 1: Load reference data
    print("\n--- Loading preview_import.xlsx (reference) ---")
    coord_keys, location_keys, preview_rows = load_preview_import()

    # Step 2: Load update.xlsx
    print("\n--- Loading update.xlsx ---")
    update_rows = load_update_sheets()

    # Step 3: Load school.xlsx (ชัยภูมิ only)
    print("\n--- Loading school.xlsx (ชัยภูมิ only) ---")
    school_rows = load_school_sheets()

    # Step 4: Find missing
    print("\n--- Finding missing rows ---")
    update_missing = find_missing(update_rows, coord_keys, location_keys, "update.xlsx")
    school_missing = find_missing(school_rows, coord_keys, location_keys, "school.xlsx")

    # Step 5: Print summaries
    print_summary(update_missing, "update.xlsx")
    print_summary(school_missing, "school.xlsx")

    # Step 6: Check overlap
    check_duplicates_between_sources(update_missing, school_missing)

    print(f"\n{'='*60}")
    print("DONE")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
