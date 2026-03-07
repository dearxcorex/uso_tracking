"""
Import USO service points using merged_update.xlsx as base + final.xlsx for lat/long & asset_id.

1. Reads merged_update.xlsx (492 service points with service_name, provider, zone, etc.)
2. Matches each to closest final.xlsx row by lat/long (within 1km)
3. Updates lat/long from final.xlsx + adds asset_id, o_asset_id
4. Clears DB and re-imports everything
"""

import math
import sys
from pathlib import Path

import openpyxl
import psycopg2


def read_env():
    env_path = Path(__file__).parent.parent / ".env"
    if not env_path.exists():
        print(f"ERROR: .env file not found at {env_path}")
        sys.exit(1)
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line.startswith("DATABASE_URL"):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    print("ERROR: DATABASE_URL not found in .env")
    sys.exit(1)


def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def load_final(xlsx_path):
    """Load unique asset_ids from final.xlsx with lat/long."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["ชัยภูมิ"]

    assets = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        aid = str(row[0]) if row[0] else None
        lat = row[35]
        lng = row[36]
        if aid and lat is not None and lng is not None and aid not in assets:
            assets[aid] = {
                "asset_id": aid,
                "o_asset_id": str(row[4]) if row[4] else None,
                "lat": float(lat),
                "lng": float(lng),
                "district": row[33],
                "subdistrict": row[32],
                "use_at": row[29],
                "use_at2": row[30],
                "ref_doc": str(row[23]) if row[23] else None,
            }
    wb.close()
    return assets


def load_merged(xlsx_path):
    """Load service points from merged_update.xlsx."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lat = float(row[8]) if row[8] is not None else None
        lng = float(row[9]) if row[9] is not None else None
        rows.append({
            "contract_number": str(row[0]) if row[0] is not None else None,
            "service_name": row[1],
            "village": row[2],
            "subdistrict": row[3],
            "district": row[4],
            "province": row[5],
            "install_location": row[6],
            "provider": row[7],
            "latitude": lat,
            "longitude": lng,
            "zone": row[10],
            "asset_id": None,
            "o_asset_id": None,
        })
    wb.close()
    return rows


def main():
    data_dir = Path(__file__).parent.parent / "data"
    final_path = data_dir / "final.xlsx"
    merged_path = data_dir / "merged_update.xlsx"

    if not final_path.exists():
        print(f"ERROR: final.xlsx not found at {final_path}")
        sys.exit(1)
    if not merged_path.exists():
        print(f"ERROR: merged_update.xlsx not found at {merged_path}")
        sys.exit(1)

    # Load both sources
    print("Loading final.xlsx...")
    final_assets = load_final(final_path)
    print(f"  {len(final_assets)} unique asset_ids with lat/long")

    print("Loading merged_update.xlsx...")
    merged_rows = load_merged(merged_path)
    print(f"  {len(merged_rows)} service points")

    # Match each merged row to closest final.xlsx asset by lat/long
    print("\nMatching merged rows to final.xlsx...")
    used_assets = set()
    matched = 0
    updated_latlong = 0
    no_match = []

    for m in merged_rows:
        if m["latitude"] is None or m["longitude"] is None:
            no_match.append(m)
            continue

        best_dist = float("inf")
        best_aid = None
        for aid, f in final_assets.items():
            d = haversine_m(m["latitude"], m["longitude"], f["lat"], f["lng"])
            if d < best_dist:
                best_dist = d
                best_aid = aid

        if best_dist < 1000:  # within 1km
            matched += 1
            f = final_assets[best_aid]
            m["asset_id"] = f["asset_id"]
            m["o_asset_id"] = f["o_asset_id"]
            # Update lat/long from final.xlsx (source of truth)
            if best_dist > 0.1:  # only count if actually different
                updated_latlong += 1
            m["latitude"] = f["lat"]
            m["longitude"] = f["lng"]
            used_assets.add(best_aid)
        else:
            no_match.append(m)

    print(f"  Matched: {matched}")
    print(f"  Lat/long updated: {updated_latlong}")
    print(f"  No match (kept original): {len(no_match)}")

    # Find new points in final.xlsx not matched to any merged row
    new_assets = {aid: f for aid, f in final_assets.items() if aid not in used_assets}
    # Group new assets by unique lat/long (they're sub-assets at same location)
    new_locations = {}
    for aid, f in new_assets.items():
        key = (round(f["lat"], 5), round(f["lng"], 5))
        if key not in new_locations:
            new_locations[key] = f

    print(f"\n  New locations in final.xlsx not in merged: {len(new_locations)}")

    # Determine service_name/provider/zone for new locations from ref_doc
    # Based on cross-referencing matched pairs:
    REF_DOC_DEFAULTS = {
        "8670223": ("Mobile", "บริษัท ทรู มูฟ เอช ยูนิเวอร์แซล คอมมิวนิเคชั่น จำกัด", "USO Zone C+"),
        "0008/2560": ("Mobile", "บริษัท ทรู มูฟ เอช ยูนิเวอร์แซล คอมมิวนิเคชั่น จำกัด", "USO Zone C+"),
        "8670265": ("ห้อง USO Wrap", "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน) (TOT เดิม)", "USO Zone C"),
        "8660259": ("Wi-Fi โรงเรียน", "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน) (TOT เดิม)", "USO Zone C+"),
    }
    FALLBACK = ("Mobile", "ไม่ระบุ", "USO Zone C+")

    if no_match:
        print(f"\n--- Unmatched merged rows (kept as-is) ---")
        for m in no_match[:10]:
            print(f"  contract={m['contract_number']}, service={m['service_name']}, "
                  f"district={m['district']}, lat={m['latitude']}, lng={m['longitude']}")
        if len(no_match) > 10:
            print(f"  ... and {len(no_match) - 10} more")

    # Import to DB
    db_url = read_env()
    print(f"\nConnecting to database...")
    conn = psycopg2.connect(db_url)
    cur = conn.cursor()

    cur.execute("DELETE FROM uso_service_point")
    deleted = cur.rowcount
    print(f"Cleared {deleted} existing rows.")

    db_cols = [
        "service_name", "village", "subdistrict", "district", "province",
        "install_location", "contract_number", "provider",
        "latitude", "longitude", "zone", "asset_id", "o_asset_id",
    ]
    placeholders = ", ".join(["%s"] * len(db_cols))
    insert_sql = f"INSERT INTO uso_service_point ({', '.join(db_cols)}) VALUES ({placeholders})"

    # Insert merged rows
    inserted = 0
    for m in merged_rows:
        values = tuple(m.get(col) for col in db_cols)
        cur.execute(insert_sql, values)
        inserted += 1

    print(f"Inserted {inserted} merged rows.")

    # Insert new locations from final.xlsx
    new_inserted = 0
    for loc in new_locations.values():
        ref = loc.get("ref_doc")
        svc, prov, zone = REF_DOC_DEFAULTS.get(ref, FALLBACK)
        values = (
            svc,
            None,  # village
            loc["subdistrict"],
            loc["district"],
            "ชัยภูมิ",
            loc.get("use_at2"),  # install_location
            None,  # contract_number
            prov,
            loc["lat"],
            loc["lng"],
            zone,
            loc["asset_id"],
            loc["o_asset_id"],
        )
        cur.execute(insert_sql, values)
        new_inserted += 1

    print(f"Inserted {new_inserted} new locations from final.xlsx.")

    conn.commit()

    # Verification
    cur.execute("SELECT COUNT(*) FROM uso_service_point")
    total = cur.fetchone()[0]
    print(f"\n--- Verification ---")
    print(f"Total rows: {total}")

    cur.execute("SELECT service_name, COUNT(*) FROM uso_service_point GROUP BY service_name ORDER BY COUNT(*) DESC")
    print("Service names:")
    for name, count in cur.fetchall():
        print(f"  {name}: {count}")

    cur.execute("SELECT COUNT(*) FROM uso_service_point WHERE asset_id IS NOT NULL")
    print(f"Rows with asset_id: {cur.fetchone()[0]}")

    cur.execute("SELECT COUNT(*) FROM uso_service_point WHERE latitude IS NULL")
    print(f"Rows with null coordinates: {cur.fetchone()[0]}")

    cur.execute("SELECT district, COUNT(*) FROM uso_service_point GROUP BY district ORDER BY district")
    print("Districts:")
    for d, c in cur.fetchall():
        print(f"  {d}: {c}")

    cur.close()
    conn.close()
    print("\nDone!")


if __name__ == "__main__":
    main()
